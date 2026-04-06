"""엑셀 생성 엔진 — openpyxl 래퍼"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import structlog

logger = structlog.get_logger(__name__)

# ── 스타일 상수 ──
HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT = Font(name="맑은 고딕", color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill("solid", fgColor="F5F7FA")
BORDER_STYLE = Border(
    bottom=Side(style="thin", color="E0E0E0"),
)
TOTAL_FONT = Font(name="맑은 고딕", bold=True, size=11)
TOTAL_FILL = PatternFill("solid", fgColor="E8EDF3")


class ExcelEngine:
    """openpyxl 기반 엑셀 생성 엔진"""

    def create(
        self,
        sheets: list[dict],
        output_path: Path,
        add_total_row: bool = True,
    ) -> dict:
        """엑셀 파일 생성

        Args:
            sheets: 시트 데이터 목록 [{"name", "headers", "rows"}]
            output_path: 저장 경로
            add_total_row: 합계 행 자동 추가

        Returns:
            dict: 생성 결과 정보
        """
        wb = Workbook()
        wb.remove(wb.active)  # 기본 시트 제거

        total_rows = 0
        for sheet_data in sheets:
            ws = wb.create_sheet(title=sheet_data.get("name", "Sheet"))
            row_count = self._write_sheet(ws, sheet_data, add_total_row)
            total_rows += row_count

        wb.save(str(output_path))
        file_size = output_path.stat().st_size

        logger.info(
            "excel_created",
            sheets=len(sheets),
            total_rows=total_rows,
            size_bytes=file_size,
            path=str(output_path),
        )

        return {
            "sheet_count": len(sheets),
            "total_rows": total_rows,
            "size_bytes": file_size,
        }

    def _write_sheet(
        self, ws, data: dict, add_total: bool
    ) -> int:
        """단일 시트 작성"""
        headers = data.get("headers", [])
        rows = data.get("rows", [])

        # ── 헤더 작성 ──
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER_STYLE

        # ── 데이터 행 작성 ──
        for row_idx, row in enumerate(rows, 2):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=self._parse_value(value))
                if fill:
                    cell.fill = fill
                cell.font = Font(name="맑은 고딕", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = BORDER_STYLE

        # ── 합계 행 ──
        if add_total and rows:
            total_row_idx = len(rows) + 2
            ws.cell(row=total_row_idx, column=1, value="합계").font = TOTAL_FONT
            ws.cell(row=total_row_idx, column=1).fill = TOTAL_FILL

            # 숫자 열에 SUM 수식 추가
            for col_idx in range(2, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                # 숫자인지 확인 후 합계
                if rows and self._is_numeric(rows[0][col_idx - 1] if col_idx - 1 < len(rows[0]) else ""):
                    formula = f"=SUM({col_letter}2:{col_letter}{len(rows) + 1})"
                    cell = ws.cell(row=total_row_idx, column=col_idx, value=formula)
                else:
                    cell = ws.cell(row=total_row_idx, column=col_idx, value="")
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL

        # ── 열 너비 자동 조정 ──
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try:
                    cell_len = len(str(cell.value or ""))
                    # 한국어는 2바이트 계산
                    korean_chars = sum(1 for c in str(cell.value or "") if ord(c) > 127)
                    cell_len += korean_chars
                    max_len = max(max_len, cell_len)
                except Exception:
                    pass
            adjusted_width = min(max_len + 4, 45)
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        # ── 행 높이 ──
        ws.row_dimensions[1].height = 30  # 헤더 행

        return len(rows)

    @staticmethod
    def _parse_value(value):
        """문자열 값을 적절한 타입으로 변환"""
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, str):
            # 숫자 변환 시도
            try:
                if "." in value:
                    return float(value)
                return int(value)
            except ValueError:
                return value
        return value

    @staticmethod
    def _is_numeric(value) -> bool:
        """값이 숫자인지 확인"""
        try:
            float(str(value).replace(",", ""))
            return True
        except (ValueError, TypeError):
            return False
